import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def tables_to_xlsx_bytes(tables: list[dict]) -> bytes:
    wb = Workbook()
    # удаляем дефолтный лист
    wb.remove(wb.active)

    for t in tables:
        page = t["page"]
        ti = t["table_index"]
        rows = t["rows"]

        title = f"p{page}_t{ti}"
        # Excel ограничение 31 символ
        ws = wb.create_sheet(title[:31])

        for r_i, row in enumerate(rows, start=1):
            for c_i, cell in enumerate(row, start=1):
                ws.cell(row=r_i, column=c_i, value=cell)

        # лёгкая авто-ширина
        if rows:
            max_cols = max(len(r) for r in rows)
            for c in range(1, max_cols + 1):
                col_letter = get_column_letter(c)
                max_len = 0
                for r in rows[:200]:  # не сканируем весь лист
                    if c-1 < len(r):
                        max_len = max(max_len, len(str(r[c-1] or "")))
                ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()