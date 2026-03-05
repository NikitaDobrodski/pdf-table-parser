from __future__ import annotations

import io
from typing import List

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ..grouping import group_tables_by_header


def build_grouped_xlsx(
    tables: List[dict],
    base_title: str = "tables",
) -> bytes:
    """
    Excel пайплайн:
      - минимальная чистка мусора
      - группировка по заголовкам
      - 1 лист = 1 формат таблицы
      - лист содержит header + все строки всех таблиц этой группы
    """
    grouped = group_tables_by_header(tables, drop_noise_rows=True)

    wb = Workbook()
    if wb.worksheets:
        wb.remove(wb.worksheets[0])

    group_no = 0
    for sig, items in grouped.items():
        group_no += 1
        title = f"g{group_no:03d}"
        if sig:
            preview = "_".join([c[:12] for c in sig[:2] if c]).strip("_")
            if preview:
                title += f"_{preview}"
        title = title[:31]

        ws = wb.create_sheet(title=title)

        header = items[0].header if items and items[0].header else []
        row_i = 1
        if header:
            for c_i, val in enumerate(header, start=1):
                ws.cell(row=row_i, column=c_i, value=val)
            row_i += 1

        for it in items:
            for r in it.rows:
                for c_i, val in enumerate(r, start=1):
                    ws.cell(row=row_i, column=c_i, value=val)
                row_i += 1

        # лёгкая авто-ширина
        max_row_scan = min(ws.max_row, 200)
        max_col = min(ws.max_column, 50)
        for c in range(1, max_col + 1):
            best = 0
            for r in range(1, max_row_scan + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    continue
                best = max(best, len(str(v)))
            ws.column_dimensions[get_column_letter(c)].width = max(8, min(60, best + 2))

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()